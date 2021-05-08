import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook(r'文件操作\2-工资处理\工资统计new.xlsx')
sheet = wb['工作量汇总']
last = sheet.max_row+1

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet['A1'] = '教师姓名'
new_sheet['B1'] = '工作量工资'


for index, values in enumerate(sheet.rows, 1):
    if 1 < index < sheet.max_row-1:
        lst = list(map(lambda x: x.value, values))
        new_sheet.cell(row=index, column=1).value = lst[1]
        new_sheet.cell(row=index, column=2).value = lst[5]

chart = BarChart()
chart.title = '工资图表'
chart.y_axis.title = '工资'
chart.x_axis.title = '姓名'
data = Reference(new_sheet, min_col=2, min_row=1, max_row=new_sheet.max_row)
cats = Reference(new_sheet, min_col=1, min_row=2, max_row=new_sheet.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
new_sheet.add_chart(chart, 'D2')

new_wb.save(r"文件操作\2-工资处理\工资图表.xlsx")
