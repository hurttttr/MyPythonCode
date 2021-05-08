import openpyxl

wb = openpyxl.load_workbook(r'文件操作\2-工资处理\工资统计.xlsx')
sheet = wb['sheet1']
sheet.title = '教工基本信息'
wb.sheetnames

sy_dic = {}
ll_dic = {}

sheet = wb['实验教学任务']
sheet1 = wb['实验工作量']

for index, values in enumerate(sheet.rows, 1):
    if index > 1:
        lst = list(map(lambda x: x.value, values))[:6]
        sheet1.cell(row=index, column=1).value = lst[4]
        sheet1.cell(row=index, column=2).value = lst[5]
        sheet1.cell(row=index, column=3).value = lst[0]
        sheet1.cell(row=index, column=4).value = lst[3]
        sheet1.cell(row=index, column=5).value = lst[1]
        sheet1.cell(row=index, column=6).value = lst[2]
        sheet1.cell(row=index, column=7).value = 0.7*(1+(lst[2]-40)*0.015)
        sheet1.cell(row=index, column=8).value = round(
            lst[1]*0.7*(1+(lst[2]-40)*0.015), 2)
        sy_dic[(lst[4], lst[5])] = sy_dic.get((lst[4], lst[5]), 0) + \
            round(lst[1]*0.7*(1+(lst[2]-40)*0.015), 2)

sheet = wb['理论教学任务']
sheet1 = wb['理论工作量']

for index, values in enumerate(sheet.rows, 1):
    if index > 1:
        lst = list(map(lambda x: x.value, values))[:7]
        ans = int(lst[1])
        sheet1.cell(row=index, column=1).value = lst[4]
        sheet1.cell(row=index, column=2).value = lst[5]
        sheet1.cell(row=index, column=3).value = lst[0]
        sheet1.cell(row=index, column=4).value = lst[1]
        sheet1.cell(row=index, column=5).value = lst[6]
        sheet1.cell(row=index, column=7).value = lst[2]
        sheet1.cell(row=index, column=9).value = lst[3]
        if lst[6] == '是':
            sheet1.cell(row=index, column=6).value = 0.9
            ans *= 0.9
        else:
            sheet1.cell(row=index, column=6).value = 1
        if lst[2] <= 50:
            sheet1.cell(row=index, column=8).value = 1
        else:
            sheet1.cell(row=index, column=8).value = 1+(lst[2]-50)/150
            ans *= 1+(lst[2]-50)/150
        sheet1.cell(row=index, column=10).value = round(ans, 2)
        ll_dic[(lst[4], lst[5])] = ll_dic.get(
            (lst[4], lst[5]), 0)+round(ans, 2)


sheet = wb['工作量汇总']
sheet1 = wb['教工基本信息']


for index, values in enumerate(sheet1.rows, 1):
    if index > 1:
        lst = list(map(lambda x: x.value, values))[:2]
        sheet.cell(row=index, column=1).value = lst[0]
        sheet.cell(row=index, column=2).value = lst[1]
        sheet.cell(row=index, column=3).value = sy_dic.get((lst[0], lst[1]), 0)
        sheet.cell(row=index, column=4).value = ll_dic.get((lst[0], lst[1]), 0)
        worktime = sy_dic.get((lst[0], lst[1]), 0) + \
            ll_dic.get((lst[0], lst[1]), 0)
        sheet.cell(row=index, column=5).value = worktime
        if worktime <= 100:
            sheet.cell(row=index, column=6).value = worktime*50
        else:
            sheet.cell(row=index, column=6).value = worktime*70


wb.save('工资统计new.xlsx')
