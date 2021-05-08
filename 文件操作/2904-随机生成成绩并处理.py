import openpyxl
import random

str = [['赵', '钱', '孙', '李'], ['伟', '昀', '琛', '东'], ['坤', '艳', '志']]
Class = ['语文', '数学', '英语']

wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = '姓名'
sheet['B1'] = '课程'
sheet['C1'] = '成绩'

for i in range(50):
    name_length = random.randint(2, 3)
    name = ''
    if name_length == 3:
        name = str[0][random.randint(
            0, 2)]+str[1][random.randint(0, 2)]+str[2][random.randint(0, 1)]
    else:
        name = str[0][random.randint(0, 2)]+str[1][random.randint(0, 2)]
    gradge = random.randint(0, 100)
    t = random.randint(0, 2)
    sheet.cell(row=i+2, column=1).value = name
    sheet.cell(row=i+2, column=2).value = Class[t]
    sheet.cell(row=i+2, column=3).value = gradge

wb.save(r"文件操作\test.xlsx")

wb1 = openpyxl.load_workbook(r"文件操作\test.xlsx")
sheet1 = wb1['Sheet']
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2['A1'] = '姓名'
sheet2['B1'] = '课程'
sheet2['C1'] = '成绩'


lst = []
for index, values in enumerate(sheet1.rows, 1):
    if index > 1:
        t = list(map(lambda x: x.value, values))
        print(t)
        lst.append(t)

lst.sort(key=lambda x: (x[0], x[1], -x[2]))
i = 0
while(i < len(lst)-1):
    if lst[i][0] == lst[i+1][0] and lst[i][1] == lst[i + 1][1]:
        lst.pop(i+1)
    else:
        i += 1

for i in range(len(lst)):
    sheet2.cell(row=i+2, column=1).value = lst[i][0]
    sheet2.cell(row=i+2, column=2).value = lst[i][1]
    sheet2.cell(row=i+2, column=3).value = lst[i][2]

wb2.save(r"文件操作\result.xlsx")
