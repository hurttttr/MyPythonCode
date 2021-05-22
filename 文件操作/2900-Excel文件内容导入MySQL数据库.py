import pymysql
import openpyxl
import os
import shutil
import random

conn = pymysql.connect(
    host='127.0.0.1',
    port=3306,
    user='root',
    password='123456',
    database='mydatabase',
    charset='utf8'
)

#创建文件夹存放excel
dir_name='excel'
if not os.path.isdir(dir_name):
    os.mkdir(dir_name)
dest = os.getcwd() + '\\' + dir_name

num = 1
for i in range(50):
    wb = openpyxl.Workbook()
    sheet = wb.active

    row = random.randint(1,5)
    for j in range(1,row+1):
        sheet.cell(row=j, column=1).value = num
        num+=1
        for k in range(2,6):
            sheet.cell(row=j, column=k).value = random.randint(0,50)

    wb.save('excel\\{}.xlsx'.format(i))

lst = []
for i in range(50):
    wb = openpyxl.load_workbook('excel\\{}.xlsx'.format(i))
    sheet = wb['Sheet']

    for index,values in enumerate(sheet.rows,1):
        tmp = tuple(map(lambda x:x.value,values))
        lst.append(tmp)

try:
    with conn.cursor() as cursor:
        sql = 'insert into excel(id,A,B,C,D) values(%s,%s,%s,%s,%s)'
        cursor.executemany(sql, lst)
        conn.commit()
except pymysql.DatabaseError:
    conn.rollback()
finally:
    conn.close()
    
#自动删除文件夹
shutil.rmtree(dest)

