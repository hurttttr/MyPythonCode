import openpyxl

wb = openpyxl.load_workbook('个人爱好.xlsx')
sheet = wb['Sheet1']
last = sheet.max_column+1
sheet.cell(row=1,column=last).value='所有爱好'
tietle = [sheet.cell(row=1,column=col).value for col in range(2,last)]

for index,value in enumerate(sheet.rows,1):
    if index!=1:
        lst = list(map(lambda x:x.value,value))[1:]
        res = [tietle[i]for i in range(len(lst)) if lst[i]=='是']
        sheet.cell(row=index,column=last).value=','.join(res)

wb.save('个人爱好.xlsx')