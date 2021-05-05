import openpyxl

judeg_dic={'A':95,'B':85,'C':75,'D':65,'E':0}

def regular_grade():
    wb = openpyxl.load_workbook('作业成绩.xlsx')
    sheet = wb['平时作业']
    last = sheet.max_column+1

    for index,value in enumerate(sheet.rows,1):
        if 4<index<sheet.max_row-1:
            lst = list(map(lambda x:x.value,value))[3:-1]
            res = 0
            for j in lst:
                res+=judeg_dic[j]
            sheet.cell(row=index,column=12).value=int(res/8+0.5)
    
    wb.save('作业成绩new.xlsx')

def test_grade():
    wb = openpyxl.load_workbook('实验成绩.xlsx')
    sheet = wb['实验报告']
    sheet1 = wb['实验成绩']
    last = sheet.max_column+1

    for index,value in enumerate(sheet.rows,1):
        if 5<index<sheet.max_row-1:
            lst = list(map(lambda x:x.value,value))[3:11]
            res = 0
            for j in lst:
                res+=judeg_dic[j]
            sheet.cell(row=index,column=12).value=int(res/8+0.5)
            sheet1.cell(row=index,column=6).value=int(res/8+0.5)

    for index,value in enumerate(sheet1.rows,1):
         if 5<index<sheet1.max_row:
            lst = list(map(lambda x:x.value,value))[3:6]
            ans = int(lst[0]*0.1+lst[1]*0.1+lst[2]*0.85+0.5)
            sheet1.cell(row=index,column=7).value=ans

    wb.save('实验成绩new.xlsx')

def cal_grade():
    wb1 = openpyxl.load_workbook('作业成绩new.xlsx')
    wb2 = openpyxl.load_workbook('实验成绩new.xlsx')
    wb3 = openpyxl.load_workbook('成绩汇总.xlsx')

    sheet1 = wb1['平时作业']
    sheet2 = wb2['实验成绩']
    sheet3 = wb3['Sheet1']

    for index,value in enumerate(sheet3.rows,1):
        if 5<index<sheet3.max_row-2:
            ps = sheet1.cell(row=index-1,column=12).value
            sy = sheet2.cell(row=index,column=7).value
            lst = list(map(lambda x:x.value,value))[5:7]
            ans = int(ps*0.1+lst[0]*0.1+lst[1]*0.6+sy*0.2+0.5)
            sheet3.cell(row=index,column=5).value=ps
            sheet3.cell(row=index,column=8).value=sy
            sheet3.cell(row=index,column=9).value=ans

    wb3.save('成绩汇总new.xlsx')


if __name__ == '__main__':
    regular_grade()
    test_grade()
    cal_grade()