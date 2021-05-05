import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

for index, values in enumerate(sheet.rows, 1):
    lst = list(map(lambda a: a.value, values))
    if lst[0] == 'Garlic':
        sheet.cell(row=index, column=2).value = 3.07
    elif lst[0] == 'Celery':
        sheet.cell(row=index, column=2).value = 1.19
    elif lst[0] == 'Lemon':
        sheet.cell(row=index, column=2).value = 1.27

wb.save('produceSales(new).xlsx')
