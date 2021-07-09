import openpyxl
import random
index = 1
str = [['赵', '钱', '孙', '李'], ['伟', '昀', '琛', '东', '凡'], ['坤', '艳', '志', '新']]
with open('sql.txt', 'w', encoding='utf-8') as f:
    wb = openpyxl.load_workbook('图书.xlsx')
    dic = {}
    pressname = []
    for i in wb:
        sheet = i
        last = sheet.max_column+1
        lst = [sheet.cell(row=1, column=col).value for col in range(1, last)]
        mlast = sheet.max_row+1
        for row in range(2, mlast):
            imformation = [sheet.cell(
                row=row, column=col).value for col in range(1, last)]
            # print(imformation)
            if imformation[2] not in pressname:
                pressname.append(imformation[2])
            dic[imformation[2]] = dic.get(
                imformation[2], []) + [imformation[1]]
            n = random.randint(1, 5)
            for j in range(n):
                sql = "insert 图书表(book_id, book_name, ISBN, author) values('{:0>4}', '{}', '{}', '{}')".format(
                    index, imformation[0], imformation[1], imformation[3])

                index += 1
                # f.write(sql)
                # f.write('\n')
    print(dic)
    press = ['高等教育出版社', '上海远东出版社', '人民教育出版社', '清华大学出版社',
             '机械工业出版社', '人民邮电出版社', '北京大学出版社', '中国人民大学出版社']

    ord = 1
    for i in press:
        ISBN_list = dic[i]
        order = '950{:0>2}'.format(ord)
        ord += 1

        for j in ISBN_list:
            sql = "insert into 订购表 values('{}','{}')".format(
                j, order,)
            print(sql)
